from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

def get_link():
  while True:
    link = input('Enter a tabroom link of the tournament results page for any prelim round: ')
    if link.startswith('https://www.tabroom.com/'):
      break
    print('Invalid link')
  return link

options = Options()
options.headless = True
driver = webdriver.Firefox(options=options)

tabroom_link = get_link()

driver.get(tabroom_link)
elements = driver.find_elements(by=By.CSS_SELECTOR, value='.marno.marvert') # this selects all the <a> elements that have links to each team's results page

team_links = []

for element in elements:
  team_links.append(element.get_attribute('href'))

# sometimes there are empty links when teams get byes, so get rid of those
for index, team_link in enumerate(team_links):
  if team_link.endswith('='):
    team_links.pop(index)

tournament_name = driver.find_element(by=By.CSS_SELECTOR, value='h2').text

print(f'\nFetching {len(team_links)} teams at tournament {tournament_name}')

teams_info = {}
team_codes = []
current_team = 0 # for numbering the list in the console output

# iterate through every team
for team_link in team_links:
  current_team += 1
  driver.get(team_link)
  team_code = driver.find_element(by=By.CSS_SELECTOR, value='h2').text
  # get names from h4 element
  names_heading = driver.find_element(by=By.CSS_SELECTOR, value='h4.nospace.semibold').text
  names = names_heading.split(' & ')
  teams_info[team_code] = {'names': names}
  # this makes sure it doesn't break if a debater goes mav
  for name in names:
    teams_info[team_code][name] = []
  # get all the elements with speaker points
  points_elements = list(driver.find_elements(by=By.CSS_SELECTOR, value='.fifth.marno'))
  if len(points_elements) == 0:
    continue

  for index, points_element in enumerate(points_elements):
    # speaker points alternate between speakers in the points_elements list
    # so if index is even, add it to speaker points list for first partner aka names[0]
    # and if index is odd, add it to speaker points list for second partner aka names[1]

    # first, handle mav debaters by just zeroing out the index and making sure it's even every time
    if len(names) == 1:
      # if mav
      index = 0
    # insert it to front instead of appending because points_elements is in reverse chronological order
    teams_info[team_code][names[index % 2]].insert(0, float(points_element.text))

  team_codes.append(team_code)
  
  print(f'\n{current_team}) {team_code}')
  for name in names:
    print(f'{name}: {teams_info[team_code][name]}')

driver.close()

print('\nSaving to a spreadsheet...')

wb = load_workbook('speakscalculatortemplate.xlsx')
ws = wb.active
active_row = 2 # start entering data on 2, increment after entering each person's points

# for every team in the tournament
for team_code in team_codes:
  # for each partner in that team code
  for name in teams_info[team_code]['names']:
    ws.cell(row=active_row, column=1, value=team_code)
    ws.cell(row=active_row, column=2, value=name)
    active_column = 3
    # for each speaker points score for that debater (in chronological order of rounds)
    for score in teams_info[team_code][name]:
      ws.cell(row=active_row, column=active_column, value=score)
      active_column += 1
    active_row += 1

filename = f"{input('File name (enter for default): ')}.xlsx"
if (filename == '.xlsx'):
  filename = f'SPEAKER AWARDS {tournament_name}.xlsx'
wb.save(filename)
print(f'Saved to {filename}')